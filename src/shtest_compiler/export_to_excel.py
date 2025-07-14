import sys
import os
from shtest_compiler.command_loader import PatternRegistry
import openpyxl
from openpyxl.utils import get_column_letter


def export_patterns_to_excel(actions_yml, validations_yml, output_xlsx):
    registry = PatternRegistry(actions_yml, validations_yml)
    wb = openpyxl.Workbook()
    ws_actions = wb.active
    ws_actions.title = "Actions"
    ws_valid = wb.create_sheet("Validations")

    # Actions
    ws_actions.append(["Phrase canonique", "Handler", "Alias"])
    for canon, entry in registry.actions["canonicals"].items():
        phrase = entry["phrase"]
        handler = entry["handler"]
        aliases = [a for a, c in registry.actions["alias_map"].items() if c == phrase]
        ws_actions.append([phrase, handler, ", ".join(aliases)])
    for col in range(1, 4):
        ws_actions.column_dimensions[get_column_letter(col)].width = 40

    # Validations
    ws_valid.append(["Phrase canonique", "Handler", "Alias"])
    for canon, entry in registry.validations["canonicals"].items():
        phrase = entry["phrase"]
        handler = entry["handler"]
        aliases = [
            a for a, c in registry.validations["alias_map"].items() if c == phrase
        ]
        ws_valid.append([phrase, handler, ", ".join(aliases)])
    for col in range(1, 4):
        ws_valid.column_dimensions[get_column_letter(col)].width = 40

    wb.save(output_xlsx)
    print(f"Exporté vers {output_xlsx}")


def export_tests_to_excel(input_dir, output_file):
    """Export patterns to Excel using default YAML files."""
    actions_yml = os.path.join(
        os.path.dirname(__file__), "config", "patterns_actions.yml"
    )
    validations_yml = os.path.join(
        os.path.dirname(__file__), "config", "patterns_validations.yml"
    )
    export_patterns_to_excel(actions_yml, validations_yml, output_file)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(
            "Usage: python export_to_excel.py <actions_yml> <validations_yml> <output_xlsx>"
        )
        sys.exit(1)
    actions_yml = sys.argv[1]
    validations_yml = sys.argv[2]
    output_xlsx = sys.argv[3]
    export_patterns_to_excel(actions_yml, validations_yml, output_xlsx)
